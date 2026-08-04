"""Admin endpoints for program discount roster (e.g. ABSN $100 off)."""

import io
import logging
import uuid
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession

from ..auth.okta import OktaUser, require_admin
from ..config import settings
from ..database import get_db
from ..models.discount_roster import DiscountRoster

logger = logging.getLogger("quarry.discount_roster")

router = APIRouter(dependencies=[Depends(require_admin())])


class RosterEntry(BaseModel):
    id: str
    student_id: str
    email: str | None
    first_name: str
    last_name: str
    program_name: str
    discount_amount: float
    academic_year: str | None
    created_at: str


class RosterAddRequest(BaseModel):
    student_id: str
    email: str | None = None
    first_name: str
    last_name: str
    program_name: str = "ABSN"
    discount_amount: float = 100
    academic_year: str | None = "2026-2027"


class RosterUploadResult(BaseModel):
    imported: int
    skipped: int
    errors: list[str]


class DiscountConfigRead(BaseModel):
    okta_groups: list[str]
    amount: float
    label: str


@router.get("/config", response_model=DiscountConfigRead)
async def get_discount_config():
    groups = [g.strip() for g in (settings.auto_discount_okta_groups or "").split(",") if g.strip()]
    return DiscountConfigRead(
        okta_groups=groups,
        amount=float(settings.auto_discount_amount or 100),
        label=settings.auto_discount_label or "ABSN Program Discount",
    )


@router.get("/roster", response_model=list[RosterEntry])
async def list_roster(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(DiscountRoster).order_by(DiscountRoster.last_name, DiscountRoster.first_name)
    )
    entries = result.scalars().all()
    return [
        RosterEntry(
            id=str(e.id),
            student_id=e.student_id,
            email=e.email,
            first_name=e.first_name,
            last_name=e.last_name,
            program_name=e.program_name,
            discount_amount=float(e.discount_amount or 0),
            academic_year=e.academic_year,
            created_at=e.created_at.isoformat() if e.created_at else "",
        )
        for e in entries
    ]


@router.get("/roster/count")
async def roster_count(db: AsyncSession = Depends(get_db)):
    count = (await db.execute(
        select(func.count()).select_from(DiscountRoster)
    )).scalar() or 0
    return {"count": count}


@router.post("/roster", response_model=RosterEntry, status_code=201)
async def add_single_entry(data: RosterAddRequest, db: AsyncSession = Depends(get_db)):
    entry = DiscountRoster(
        student_id=data.student_id.strip(),
        email=data.email.strip() if data.email else None,
        first_name=data.first_name.strip(),
        last_name=data.last_name.strip(),
        program_name=data.program_name.strip() or "ABSN",
        discount_amount=Decimal(str(data.discount_amount)),
        academic_year=data.academic_year,
    )
    db.add(entry)
    await db.flush()
    await db.refresh(entry)
    return RosterEntry(
        id=str(entry.id),
        student_id=entry.student_id,
        email=entry.email,
        first_name=entry.first_name,
        last_name=entry.last_name,
        program_name=entry.program_name,
        discount_amount=float(entry.discount_amount or 0),
        academic_year=entry.academic_year,
        created_at=entry.created_at.isoformat() if entry.created_at else "",
    )


@router.post("/roster/upload", response_model=RosterUploadResult)
async def upload_roster(
    file: UploadFile = File(...),
    program_name: str = Form("ABSN"),
    discount_amount: float = Form(100),
    academic_year: str = Form("2026-2027"),
    replace: bool = Form(False),
    db: AsyncSession = Depends(get_db),
    _admin: OktaUser = Depends(require_admin()),
):
    """Upload Excel/CSV. Columns: student_id (Moravian ID), last, first, email (optional)."""
    content = await file.read()
    filename = file.filename or ""

    rows: list[dict] = []
    errors: list[str] = []

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers_raw = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = _normalize_headers(headers_raw)
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row_cells)))
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel file: {e}")
    elif filename.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        raw_fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
        norm_fieldnames = _normalize_headers(raw_fieldnames)
        for raw_row in reader:
            row_dict = {
                norm_fieldnames[i]: v
                for i, (_, v) in enumerate(raw_row.items())
                if i < len(norm_fieldnames)
            }
            rows.append(row_dict)
    else:
        raise HTTPException(400, "Upload an .xlsx or .csv file")

    if not rows:
        raise HTTPException(400, "File contains no data rows")

    if replace:
        await db.execute(delete(DiscountRoster).where(DiscountRoster.program_name == program_name))

    amount = Decimal(str(discount_amount))
    imported = 0
    skipped = 0
    for i, row in enumerate(rows, start=2):
        student_id = str(row.get("student_id") or "").strip()
        if not student_id or student_id.lower() == "none":
            errors.append(f"Row {i}: missing student ID")
            skipped += 1
            continue

        entry = DiscountRoster(
            student_id=student_id.split("@")[0],
            email=(str(row.get("email") or "").strip() or None),
            first_name=str(row.get("first_name") or "").strip(),
            last_name=str(row.get("last_name") or "").strip(),
            program_name=program_name,
            discount_amount=amount,
            academic_year=academic_year,
        )
        db.add(entry)
        imported += 1

    if imported > 0:
        await db.flush()

    return RosterUploadResult(imported=imported, skipped=skipped, errors=errors[:20])


@router.delete("/roster/{entry_id}")
async def delete_roster_entry(entry_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    entry = await db.get(DiscountRoster, entry_id)
    if not entry:
        raise HTTPException(404, "Entry not found")
    await db.delete(entry)
    await db.flush()
    return {"deleted": True}


@router.delete("/roster")
async def clear_roster(
    program_name: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    if program_name:
        count = (await db.execute(
            select(func.count()).select_from(DiscountRoster).where(
                DiscountRoster.program_name == program_name
            )
        )).scalar() or 0
        await db.execute(delete(DiscountRoster).where(DiscountRoster.program_name == program_name))
    else:
        count = (await db.execute(
            select(func.count()).select_from(DiscountRoster)
        )).scalar() or 0
        await db.execute(delete(DiscountRoster))
    return {"deleted": count}


def _normalize_headers(headers: list[str]) -> list[str]:
    mapping = {
        "moravian id": "student_id",
        "moravian_id": "student_id",
        "id": "student_id",
        "student id": "student_id",
        "student_id": "student_id",
        "last": "last_name",
        "last_name": "last_name",
        "last name": "last_name",
        "first": "first_name",
        "first_name": "first_name",
        "first name": "first_name",
        "email": "email",
        "program": "program_name",
        "program_name": "program_name",
    }
    return [mapping.get(h, h) for h in headers]
