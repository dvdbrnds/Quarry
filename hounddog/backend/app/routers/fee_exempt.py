"""Admin endpoints for managing the fee-exempt roster."""

import io
import logging
import uuid
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession

from ..auth.okta import OktaUser, require_admin
from ..config import settings
from ..database import get_db
from ..models.fee_exempt_roster import FeeExemptRoster
from ..models.lottery_v2 import LotteryV2Application
from ..models.payment import Payment
from ..models.permit_type import PermitType
from ..services.roster_permit_status import (
    PermitMatch,
    load_active_permit_indexes,
    match_roster_to_permit,
)

logger = logging.getLogger("quarry.fee_exempt")

router = APIRouter(dependencies=[Depends(require_admin())])


class RosterEntry(BaseModel):
    id: str
    student_id: str
    email: str | None
    first_name: str
    last_name: str
    reason: str
    building: str | None
    room: str | None
    academic_year: str | None
    created_at: str
    has_permit: bool = False
    permit_number: str | None = None
    permit_type: str | None = None
    matched_by: str | None = None


class RosterAddRequest(BaseModel):
    student_id: str
    email: str | None = None
    first_name: str
    last_name: str
    reason: str = "Res Life Staff"
    building: str | None = None
    room: str | None = None
    academic_year: str | None = "2026-2027"


class RosterUploadResult(BaseModel):
    imported: int
    skipped: int
    errors: list[str]


def _entry_from_row(e: FeeExemptRoster, match: PermitMatch | None = None) -> RosterEntry:
    m = match or PermitMatch()
    return RosterEntry(
        id=str(e.id),
        student_id=e.student_id,
        email=e.email,
        first_name=e.first_name,
        last_name=e.last_name,
        reason=e.reason,
        building=e.building,
        room=e.room,
        academic_year=e.academic_year,
        created_at=e.created_at.isoformat() if e.created_at else "",
        has_permit=m.has_permit,
        permit_number=m.permit_number,
        permit_type=m.permit_type,
        matched_by=m.matched_by,
    )


@router.get("/roster", response_model=list[RosterEntry])
async def list_roster(db: AsyncSession = Depends(get_db)):
    """List all fee-exempt students with active-permit status."""
    result = await db.execute(
        select(FeeExemptRoster).order_by(FeeExemptRoster.last_name, FeeExemptRoster.first_name)
    )
    entries = result.scalars().all()
    by_email, by_student_id, by_name = await load_active_permit_indexes(db)
    return [
        _entry_from_row(
            e,
            match_roster_to_permit(
                student_id=e.student_id,
                email=e.email,
                first_name=e.first_name,
                last_name=e.last_name,
                by_email=by_email,
                by_student_id=by_student_id,
                by_name=by_name,
            ),
        )
        for e in entries
    ]


@router.get("/roster/count")
async def roster_count(db: AsyncSession = Depends(get_db)):
    count = (await db.execute(
        select(func.count()).select_from(FeeExemptRoster)
    )).scalar() or 0
    return {"count": count}


@router.post("/roster", response_model=RosterEntry, status_code=201)
async def add_single_entry(data: RosterAddRequest, db: AsyncSession = Depends(get_db)):
    """Add a single person to the fee-exempt roster."""
    entry = FeeExemptRoster(
        student_id=data.student_id.strip(),
        email=data.email.strip() if data.email else None,
        first_name=data.first_name.strip(),
        last_name=data.last_name.strip(),
        reason=data.reason,
        building=data.building,
        room=data.room,
        academic_year=data.academic_year,
    )
    db.add(entry)
    await db.flush()
    await db.refresh(entry)
    by_email, by_student_id, by_name = await load_active_permit_indexes(db)
    return _entry_from_row(
        entry,
        match_roster_to_permit(
            student_id=entry.student_id,
            email=entry.email,
            first_name=entry.first_name,
            last_name=entry.last_name,
            by_email=by_email,
            by_student_id=by_student_id,
            by_name=by_name,
        ),
    )


@router.post("/roster/upload", response_model=RosterUploadResult)
async def upload_roster(
    file: UploadFile = File(...),
    reason: str = Form("Res Life Staff"),
    academic_year: str = Form("2026-2027"),
    replace: bool = Form(False),
    db: AsyncSession = Depends(get_db),
    _admin: OktaUser = Depends(require_admin()),
):
    """Upload an Excel or CSV file to populate the fee-exempt roster.

    Expected columns: student_id (or moravian_id/id), last (or last_name),
    first (or first_name), building (optional), room (optional), email (optional).
    """
    content = await file.read()
    filename = file.filename or ""

    rows: list[dict] = []
    errors: list[str] = []

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers_raw = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            headers = _normalize_headers(headers_raw)
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row_cells))
                rows.append(row_dict)
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel file: {e}")
    elif filename.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        raw_fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
        norm_fieldnames = _normalize_headers(raw_fieldnames)
        for raw_row in reader:
            row_dict = {
                norm_fieldnames[i]: v
                for i, (_, v) in enumerate(raw_row.items())
                if i < len(norm_fieldnames)
            }
            rows.append(row_dict)
    else:
        raise HTTPException(400, "Upload an .xlsx or .csv file")

    if not rows:
        raise HTTPException(400, "File contains no data rows")

    if replace:
        await db.execute(delete(FeeExemptRoster))

    imported = 0
    skipped = 0
    for i, row in enumerate(rows, start=2):
        student_id = str(row.get("student_id") or "").strip()
        if not student_id:
            errors.append(f"Row {i}: missing student ID")
            skipped += 1
            continue
        student_id = student_id.split("@")[0]

        first_name = str(row.get("first_name") or "").strip()
        last_name = str(row.get("last_name") or "").strip()
        email = str(row.get("email") or "").strip() or None
        building = str(row.get("building") or "").strip() or None
        room = str(row.get("room") or "").strip() or None

        entry = FeeExemptRoster(
            student_id=student_id,
            email=email,
            first_name=first_name,
            last_name=last_name,
            reason=reason,
            building=building,
            room=room,
            academic_year=academic_year,
        )
        db.add(entry)
        imported += 1

    if imported > 0:
        await db.flush()

    return RosterUploadResult(imported=imported, skipped=skipped, errors=errors[:20])


@router.delete("/roster/{entry_id}")
async def delete_roster_entry(
    entry_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    _admin: OktaUser = Depends(require_admin()),
):
    entry = await db.get(FeeExemptRoster, entry_id)
    if not entry:
        raise HTTPException(404, "Entry not found")
    await db.delete(entry)
    await db.flush()
    return {"deleted": True}


class RosterUpdateRequest(BaseModel):
    email: str | None = None
    first_name: str | None = None
    last_name: str | None = None
    student_id: str | None = None
    building: str | None = None
    room: str | None = None
    reason: str | None = None


@router.patch("/roster/{entry_id}", response_model=RosterEntry)
async def update_roster_entry(
    entry_id: uuid.UUID,
    data: RosterUpdateRequest,
    db: AsyncSession = Depends(get_db),
    _admin: OktaUser = Depends(require_admin()),
):
    """Update a fee-exempt roster row (e.g. backfill email for matching)."""
    entry = await db.get(FeeExemptRoster, entry_id)
    if not entry:
        raise HTTPException(404, "Entry not found")
    if data.email is not None:
        entry.email = data.email.strip().lower() or None
    if data.first_name is not None:
        entry.first_name = data.first_name.strip()
    if data.last_name is not None:
        entry.last_name = data.last_name.strip()
    if data.student_id is not None:
        entry.student_id = data.student_id.strip()
    if data.building is not None:
        entry.building = data.building.strip() or None
    if data.room is not None:
        entry.room = data.room.strip() or None
    if data.reason is not None:
        entry.reason = data.reason.strip() or entry.reason
    await db.flush()
    await db.refresh(entry)
    by_email, by_student_id, by_name = await load_active_permit_indexes(db)
    return _entry_from_row(
        entry,
        match_roster_to_permit(
            student_id=entry.student_id,
            email=entry.email,
            first_name=entry.first_name,
            last_name=entry.last_name,
            by_email=by_email,
            by_student_id=by_student_id,
            by_name=by_name,
        ),
    )


@router.delete("/roster")
async def clear_roster(
    db: AsyncSession = Depends(get_db),
    _admin: OktaUser = Depends(require_admin()),
):
    """Delete all entries from the fee-exempt roster."""
    count = (await db.execute(
        select(func.count()).select_from(FeeExemptRoster)
    )).scalar() or 0
    await db.execute(delete(FeeExemptRoster))
    return {"deleted": count}


def _normalize_headers(headers: list[str]) -> list[str]:
    """Map common header variations to canonical field names."""
    mapping = {
        "moravian id": "student_id",
        "moravian_id": "student_id",
        "id": "student_id",
        "student id": "student_id",
        "student_id": "student_id",
        "last": "last_name",
        "last_name": "last_name",
        "last name": "last_name",
        "first": "first_name",
        "first_name": "first_name",
        "first name": "first_name",
        "email": "email",
        "building": "building",
        "hall": "building",
        "residence hall": "building",
        "room": "room",
        "room #": "room",
    }
    return [mapping.get(h, h) for h in headers]


# ── Balance Due (RA discount difference collection) ──────────────────────


class BalanceDueRow(BaseModel):
    application_id: str
    student_name: str
    email: str
    permit_type: str
    list_price: str
    expected_price: str
    amount_paid: str
    balance_due: str
    permit_number: str | None = None
    payment_link_sent: bool = False


class BalanceDueResponse(BaseModel):
    rows: list[BalanceDueRow]
    total_owed: str
    count: int


@router.get("/balance-due", response_model=BalanceDueResponse)
async def get_balance_due(db: AsyncSession = Depends(get_db)):
    """RAs who checked out under the old pricing and owe a balance."""
    ra_discount = Decimal(str(settings.ra_discount_amount))

    # Find accepted lottery apps that are fee_exempt (RA roster match)
    apps_result = await db.execute(
        select(LotteryV2Application).where(
            LotteryV2Application.fee_exempt == True,  # noqa: E712
            LotteryV2Application.status == "accepted",
        )
    )
    apps = apps_result.scalars().all()
    if not apps:
        return BalanceDueResponse(rows=[], total_owed="0.00", count=0)

    # Load permit types for price lookup
    pt_ids = {a.assigned_permit_type_id for a in apps if a.assigned_permit_type_id}
    pt_map: dict[uuid.UUID, PermitType] = {}
    if pt_ids:
        pt_result = await db.execute(select(PermitType).where(PermitType.id.in_(pt_ids)))
        for pt in pt_result.scalars().all():
            pt_map[pt.id] = pt

    # Load payments for these students (lottery_v2_permit type)
    emails = {a.student_email.lower() for a in apps if a.student_email}
    payment_by_email: dict[str, Payment] = {}
    if emails:
        pay_result = await db.execute(
            select(Payment).where(
                func.lower(Payment.payer_email).in_(emails),
                Payment.payment_type.in_(["lottery_v2_permit", "lottery_permit"]),
            )
        )
        for p in pay_result.scalars().all():
            key = (p.payer_email or "").lower()
            if key not in payment_by_email or p.amount > payment_by_email[key].amount:
                payment_by_email[key] = p

    # Load permits for permit_number lookup
    from ..models.permit import Permit
    permit_by_email: dict[str, Permit] = {}
    if emails:
        perm_result = await db.execute(
            select(Permit).where(
                func.lower(Permit.email).in_(emails),
                Permit.status == "active",
            )
        )
        for pm in perm_result.scalars().all():
            permit_by_email[(pm.email or "").lower()] = pm

    rows: list[BalanceDueRow] = []
    total = Decimal("0.00")

    for app in apps:
        pt = pt_map.get(app.assigned_permit_type_id) if app.assigned_permit_type_id else None
        if not pt:
            continue

        expected_price = max(Decimal("0.00"), pt.price - ra_discount)
        email_key = (app.student_email or "").lower()
        payment = payment_by_email.get(email_key)
        amount_paid = payment.amount if payment else Decimal("0.00")
        balance = expected_price - amount_paid

        if balance <= 0:
            continue

        permit = permit_by_email.get(email_key)
        payment_link_sent = bool(
            app.admin_notes and "balance_due_session" in app.admin_notes
        )

        rows.append(BalanceDueRow(
            application_id=str(app.id),
            student_name=app.student_name,
            email=app.student_email,
            permit_type=pt.label,
            list_price=str(pt.price),
            expected_price=str(expected_price),
            amount_paid=str(amount_paid),
            balance_due=str(balance),
            permit_number=permit.permit_number if permit else None,
            payment_link_sent=payment_link_sent,
        ))
        total += balance

    return BalanceDueResponse(
        rows=rows,
        total_owed=str(total),
        count=len(rows),
    )


@router.post("/balance-due/{app_id}/send-payment")
async def send_balance_payment(
    app_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    admin: OktaUser = Depends(require_admin()),
):
    """Create a Stripe Checkout Session for the RA balance and email the link."""
    ra_discount = Decimal(str(settings.ra_discount_amount))

    app = await db.get(LotteryV2Application, app_id)
    if not app:
        raise HTTPException(404, "Application not found")
    if not app.fee_exempt or app.status != "accepted":
        raise HTTPException(400, "Application is not an accepted RA entry")

    pt = await db.get(PermitType, app.assigned_permit_type_id) if app.assigned_permit_type_id else None
    if not pt:
        raise HTTPException(404, "Permit type not found")

    expected_price = max(Decimal("0.00"), pt.price - ra_discount)

    # Find what they actually paid
    pay_result = await db.execute(
        select(Payment).where(
            func.lower(Payment.payer_email) == app.student_email.lower(),
            Payment.payment_type.in_(["lottery_v2_permit", "lottery_permit"]),
        ).order_by(Payment.amount.desc()).limit(1)
    )
    payment = pay_result.scalar()
    amount_paid = payment.amount if payment else Decimal("0.00")
    balance = expected_price - amount_paid

    if balance <= 0:
        raise HTTPException(400, "No balance due for this student")

    if not settings.stripe_secret_key:
        raise HTTPException(503, "Stripe not configured")

    import stripe
    stripe.api_key = settings.stripe_secret_key

    base_url = settings.cors_origins[0] if settings.cors_origins else "http://localhost:5173"

    session = stripe.checkout.Session.create(
        customer_email=app.student_email,
        line_items=[{
            "price_data": {
                "currency": "usd",
                "product_data": {
                    "name": f"RA Parking Permit — Balance Due",
                    "description": (
                        f"{pt.label} permit balance: ${pt.price:.2f} list "
                        f"− ${ra_discount:.0f} RA discount "
                        f"− ${amount_paid:.2f} paid = ${balance:.2f} due"
                    ),
                },
                "unit_amount": int(balance * 100),
            },
            "quantity": 1,
        }],
        mode="payment",
        payment_intent_data={
            "statement_descriptor_suffix": "PARK PERMIT",
            "metadata": {
                "type": "ra_balance_due",
                "revenue_category": "parking_permits",
                "department": "parking_services",
                "gl_string": settings.gl_segment_separator.join([
                    settings.gl_fund, settings.gl_org,
                    settings.gl_account_permits, settings.gl_activity_permits,
                    settings.gl_segment5, settings.gl_segment6,
                ]),
                "gl_fund": settings.gl_fund,
                "gl_org": settings.gl_org,
                "gl_account": settings.gl_account_permits,
                "gl_activity": settings.gl_activity_permits,
                "permit_type_code": pt.code,
                "permit_type_label": pt.label,
                "application_id": str(app.id),
                "student_name": app.student_name,
                "student_email": app.student_email,
                "balance_amount": str(balance),
                "institution": settings.school_name or "moravian",
            },
        },
        success_url=f"{base_url}/parking?payment=success&session_id={{CHECKOUT_SESSION_ID}}",
        cancel_url=f"{base_url}/parking?payment=cancelled",
        metadata={
            "type": "ra_balance_due",
            "application_id": str(app.id),
            "permit_type_code": pt.code,
            "student_name": app.student_name,
            "email": app.student_email,
        },
    )

    # Track that we sent a payment link via admin_notes
    notes = app.admin_notes or ""
    app.admin_notes = f"{notes}\nbalance_due_session:{session.id}".strip()
    await db.flush()

    # Best-effort email
    from ..models.permit import Permit
    perm_result = await db.execute(
        select(Permit).where(
            func.lower(Permit.email) == app.student_email.lower(),
            Permit.status == "active",
        ).limit(1)
    )
    permit = perm_result.scalar()
    permit_number = permit.permit_number if permit else "N/A"

    try:
        from ..services.email import send_payment_link_email
        await send_payment_link_email(
            recipient_email=app.student_email,
            recipient_name=app.student_name,
            permit_type_label=pt.label,
            permit_number=permit_number,
            amount_display=f"${balance:.2f}",
            checkout_url=session.url,
        )
    except Exception:
        logger.exception("Failed to send balance-due email to %s", app.student_email)

    return {
        "checkout_url": session.url,
        "session_id": session.id,
        "balance_due": str(balance),
        "email_sent": True,
    }


# ── Refund Due (RAs who paid full price without their discount) ──────


class RefundDueRow(BaseModel):
    application_id: str  # permit ID (or app ID for legacy)
    student_name: str
    email: str
    permit_type: str
    list_price: str
    expected_price: str
    amount_paid: str
    refund_amount: str
    permit_number: str | None = None
    stripe_payment_id: str | None = None
    refund_issued: bool = False


class RefundDueResponse(BaseModel):
    rows: list[RefundDueRow]
    total_refundable: str
    count: int
    debug: dict | None = None


@router.get("/refund-due")
async def get_refund_due(db: AsyncSession = Depends(get_db)):
    """RAs who paid more than (list_price - $50 discount). Works from permits + payments."""
    from ..models.fee_exempt_roster import FeeExemptRoster
    from ..models.permit import Permit

    ra_discount = Decimal(str(settings.ra_discount_amount))

    # Get all RA identifiers from the roster
    roster_result = await db.execute(
        select(
            FeeExemptRoster.email,
            FeeExemptRoster.student_id,
            FeeExemptRoster.first_name,
            FeeExemptRoster.last_name,
        )
    )
    roster_rows = roster_result.all()
    ra_emails = {(r[0] or "").lower() for r in roster_rows if r[0]}
    ra_student_ids = {r[1] for r in roster_rows if r[1]}
    # Build name set for fallback matching (lowercase "first last")
    ra_names = {
        f"{(r[2] or '').strip()} {(r[3] or '').strip()}".lower()
        for r in roster_rows
        if r[2] and r[3]
    }
    # Also build first/last pairs for fuzzy matching (permit name may have middle name)
    ra_first_last = [
        ((r[2] or "").strip().lower(), (r[3] or "").strip().lower())
        for r in roster_rows
        if r[2] and r[3]
    ]

    if not ra_emails and not ra_student_ids and not ra_names:
        return RefundDueResponse(rows=[], total_refundable="0.00", count=0)

    # Find ALL active student permits (we'll filter RA membership in Python
    # because email/student_id formats may not match between roster and permits)
    perm_result = await db.execute(
        select(Permit).where(
            Permit.status == "active",
            Permit.deleted_at.is_(None),
        )
    )
    all_permits = perm_result.scalars().all()

    # Filter to only RA permits: match by email, student_id, name, or email username
    def _name_matches_roster(permit_name: str) -> bool:
        pn = permit_name.lower().strip()
        if not pn:
            return False
        if pn in ra_names:
            return True
        for first, last in ra_first_last:
            if first and last and first in pn and last in pn:
                return True
        return False

    def _email_matches_roster(email: str) -> bool:
        """Check if email username matches 'lastname + first_initial' pattern."""
        if not email or "@" not in email:
            return False
        username = email.split("@")[0].lower().rstrip("0123456789")
        for first, last in ra_first_last:
            if not last or not first:
                continue
            # Match patterns like "davisk" = "davis" + "k"(atelyn)
            candidate = last + first[0]
            if username == candidate:
                return True
        return False

    permits = []
    for p in all_permits:
        p_email = (p.email or "").lower()
        p_sid = (p.student_id or "").strip()
        if p_email in ra_emails:
            permits.append(p)
        elif p_sid in ra_student_ids:
            permits.append(p)
        elif _name_matches_roster(p.name or ""):
            permits.append(p)
        elif _email_matches_roster(p_email):
            permits.append(p)

    if not permits:
        return RefundDueResponse(rows=[], total_refundable="0.00", count=0)

    # Expand ra_emails to include permit emails found via other matches
    for p in permits:
        if p.email:
            ra_emails.add(p.email.lower())

    # Resolve permit types by code
    pt_codes = {p.permit_type for p in permits if p.permit_type}
    pt_by_code: dict[str, PermitType] = {}
    if pt_codes:
        pt_result = await db.execute(
            select(PermitType).where(PermitType.code.in_(pt_codes))
        )
        for pt in pt_result.scalars().all():
            pt_by_code[pt.code] = pt

    # Find permit purchase payments — broad search by email, plate, or stripe session
    permit_types_for_pay = [
        "lottery_v2_permit", "lottery_permit",
        "permit_purchase", "standalone_permit_purchase",
        "direct_permit_purchase", "admin_permit_charge",
    ]
    payment_by_email: dict[str, Payment] = {}

    # Search by email
    pay_result = await db.execute(
        select(Payment).where(
            func.lower(Payment.payer_email).in_(ra_emails),
            Payment.payment_type.in_(permit_types_for_pay),
        )
    )
    for p in pay_result.scalars().all():
        key = (p.payer_email or "").lower()
        if key not in payment_by_email or p.amount > payment_by_email[key].amount:
            payment_by_email[key] = p

    # Also search by plate for permits with no email-matched payment
    all_plates = {plate for perm in permits for plate in (perm.plates or [])}
    if all_plates:
        plate_pay_result = await db.execute(
            select(Payment).where(
                Payment.plate.in_(all_plates),
                Payment.payment_type.in_(permit_types_for_pay),
            )
        )
        plate_payment_map: dict[str, Payment] = {}
        for p in plate_pay_result.scalars().all():
            if p.plate:
                plate_payment_map[p.plate] = p

    # Also try matching by stripe_session_id on the permit
    stripe_sessions = {perm.stripe_session_id for perm in permits if perm.stripe_session_id}
    session_payment_map: dict[str, Payment] = {}
    if stripe_sessions:
        sess_pay_result = await db.execute(
            select(Payment).where(
                Payment.stripe_payment_id.in_(stripe_sessions),
            )
        )
        for p in sess_pay_result.scalars().all():
            if p.stripe_payment_id:
                session_payment_map[p.stripe_payment_id] = p

    # Check for refund notes on lottery apps (if any exist)
    refund_notes: dict[str, bool] = {}
    apps_result = await db.execute(
        select(LotteryV2Application.student_email, LotteryV2Application.admin_notes).where(
            func.lower(LotteryV2Application.student_email).in_(ra_emails),
        )
    )
    for email, notes in apps_result.all():
        if notes and "refund_issued" in notes:
            refund_notes[(email or "").lower()] = True

    rows: list[RefundDueRow] = []
    total = Decimal("0.00")
    seen_emails: set[str] = set()

    for permit in permits:
        email_key = (permit.email or "").lower()
        if email_key in seen_emails:
            continue

        pt = pt_by_code.get(permit.permit_type)
        if not pt:
            continue

        expected_price = max(Decimal("0.00"), pt.price - ra_discount)
        # Try email match, then plate match, then stripe session match
        payment = payment_by_email.get(email_key)
        if not payment:
            for plate in (permit.plates or []):
                payment = plate_payment_map.get(plate)
                if payment:
                    break
        if not payment and permit.stripe_session_id:
            payment = session_payment_map.get(permit.stripe_session_id)

        # If no Payment record found but permit exists with a stripe session,
        # the RA paid the full list price (permit wouldn't exist otherwise)
        if payment:
            amount_paid = payment.amount
            stripe_id = payment.stripe_payment_id
        elif permit.stripe_session_id:
            amount_paid = pt.price
            stripe_id = permit.stripe_session_id
        else:
            continue

        overpaid = amount_paid - expected_price
        if overpaid <= 0:
            continue

        seen_emails.add(email_key)

        rows.append(RefundDueRow(
            application_id=str(permit.id),
            student_name=permit.name,
            email=permit.email or "",
            permit_type=pt.label,
            list_price=str(pt.price),
            expected_price=str(expected_price),
            amount_paid=str(amount_paid),
            refund_amount=str(overpaid),
            permit_number=permit.permit_number,
            stripe_payment_id=stripe_id,
            refund_issued=bool(permit.refund_id) or refund_notes.get(email_key, False),
        ))
        total += overpaid

    # Debug: trace every RA permit and why it was included or skipped
    debug_permits = []
    for permit in permits:
        ek = (permit.email or "").lower()
        pt = pt_by_code.get(permit.permit_type)
        pay = payment_by_email.get(ek)
        plate_pay = None
        for plate in (permit.plates or []):
            plate_pay = plate_payment_map.get(plate)
            if plate_pay:
                break
        sess_pay = session_payment_map.get(permit.stripe_session_id) if permit.stripe_session_id else None
        debug_permits.append({
            "name": permit.name,
            "email": permit.email,
            "student_id": permit.student_id,
            "permit_type": permit.permit_type,
            "permit_number": permit.permit_number,
            "stripe_session_id": permit.stripe_session_id,
            "pt_found": pt is not None,
            "pt_price": str(pt.price) if pt else None,
            "payment_by_email": bool(pay),
            "payment_by_plate": bool(plate_pay),
            "payment_by_session": bool(sess_pay),
            "pay_amount": str(pay.amount) if pay else (str(plate_pay.amount) if plate_pay else (str(sess_pay.amount) if sess_pay else None)),
        })

    # Build roster detail for debug
    roster_detail = [
        {
            "first": r[2], "last": r[3],
            "email": r[0], "sid": r[1],
            "match_name": f"{(r[2] or '').strip()} {(r[3] or '').strip()}".lower(),
        }
        for r in roster_rows
    ]

    # Direct search for "davis" in all permits and payments
    davis_permits = [
        {"name": p.name, "email": p.email, "student_id": p.student_id,
         "permit_type": p.permit_type, "permit_number": p.permit_number,
         "status": p.status, "plates": p.plates}
        for p in all_permits
        if "davis" in (p.name or "").lower() or "davisk" in (p.email or "").lower()
    ]
    davis_payments_result = await db.execute(
        select(Payment).where(
            func.lower(Payment.payer_email).like("%davis%")
            | func.lower(Payment.payer_email).like("%davisk%")
        )
    )
    davis_payments = [
        {"email": p.payer_email, "amount": str(p.amount), "type": p.payment_type,
         "plate": p.plate, "stripe_id": p.stripe_payment_id, "desc": p.description}
        for p in davis_payments_result.scalars().all()
    ]

    return RefundDueResponse(
        rows=rows,
        debug={
            "davis_permits": davis_permits,
            "davis_payments": davis_payments,
            "permits_found": len(permits),
            "all_active_permits_checked": len(all_permits),
            "payments_found": len(payment_by_email),
        },
        total_refundable=str(total),
        count=len(rows),
    )


@router.post("/refund-due/{permit_id}/issue-refund")
async def issue_refund(
    permit_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    admin: OktaUser = Depends(require_admin()),
):
    """Issue a partial Stripe refund for the RA discount amount."""
    from ..models.permit import Permit

    ra_discount = Decimal(str(settings.ra_discount_amount))

    permit = await db.get(Permit, permit_id)
    if not permit:
        raise HTTPException(404, "Permit not found")

    pt_result = await db.execute(
        select(PermitType).where(PermitType.code == permit.permit_type)
    )
    pt = pt_result.scalars().first()
    if not pt:
        raise HTTPException(404, "Permit type not found")

    expected_price = max(Decimal("0.00"), pt.price - ra_discount)

    pay_result = await db.execute(
        select(Payment).where(
            func.lower(Payment.payer_email) == (permit.email or "").lower(),
            Payment.payment_type.in_([
                "lottery_v2_permit", "lottery_permit",
                "permit_purchase", "standalone_permit_purchase",
                "direct_permit_purchase", "admin_permit_charge",
            ]),
        ).order_by(Payment.amount.desc()).limit(1)
    )
    payment = pay_result.scalar()

    if payment:
        amount_paid = payment.amount
        stripe_id = payment.stripe_payment_id
    elif permit.stripe_session_id:
        amount_paid = pt.price
        stripe_id = permit.stripe_session_id
    else:
        raise HTTPException(400, "No payment found for this student")

    overpaid = amount_paid - expected_price
    if overpaid <= 0:
        raise HTTPException(400, "Student did not overpay — no refund needed")

    if not stripe_id:
        raise HTTPException(400, "No Stripe payment ID on record — manual refund required")

    if not settings.stripe_secret_key:
        raise HTTPException(503, "Stripe not configured")

    import stripe
    stripe.api_key = settings.stripe_secret_key

    try:
        if stripe_id.startswith("cs_"):
            session = stripe.checkout.Session.retrieve(stripe_id)
            stripe_id = session.payment_intent

        refund = stripe.Refund.create(
            payment_intent=stripe_id,
            amount=int(overpaid * 100),
            reason="requested_by_customer",
            metadata={
                "type": "ra_discount_refund",
                "permit_id": str(permit.id),
                "student_email": permit.email,
                "refund_amount": str(overpaid),
                "admin": admin.email or admin.sub,
            },
        )
    except stripe.StripeError as e:
        raise HTTPException(400, f"Stripe refund failed: {e.user_message or str(e)}")

    permit.refund_id = refund.id

    # Also mark refund on lottery app if one exists (legacy tracking)
    app_result = await db.execute(
        select(LotteryV2Application).where(
            func.lower(LotteryV2Application.student_email) == (permit.email or "").lower(),
        ).limit(1)
    )
    app = app_result.scalars().first()
    if app:
        notes = app.admin_notes or ""
        app.admin_notes = f"{notes}\nrefund_issued:{refund.id} ${overpaid:.2f} by {admin.email}".strip()

    await db.flush()

    logger.info(
        "Refund %s issued for %s (%s): $%s by %s",
        refund.id, permit.name, permit.email, overpaid, admin.email,
    )

    return {
        "refund_id": refund.id,
        "refund_amount": str(overpaid),
        "status": refund.status,
    }
